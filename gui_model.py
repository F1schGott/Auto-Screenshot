import os
import os.path
import time
import tkinter as tk
import tkinter.messagebox
from tkinter import filedialog
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook


class GUI:
    def __init__(self, root):
        self.root = root
        self.name_list = ''
        self.url_list = ''
        self.filepath = ''
        self.folderpath = ''
        self.success_count = 0
        self.fail_count = 0
        self.fail_list = []
        self.check = False

    def set(self):
        self.root.title("小红书自动截图工具")
        self.root.geometry('600x350+400+200')
        # 标签
        self.list_label = tk.Label(self.root, text="截图名单: ")
        self.list_label.grid(row=0, column=0)
        self.result_label = tk.Label(self.root, text="输出目录: ")
        self.result_label.grid(row=1, column=0)
        self.info_label = tk.Label(self.root, text="截图情况:")
        self.info_label.grid(row=10, column=0)
        self.width_label = tk.Label(self.root, text="图片宽度:")
        self.width_label.grid(row=0, column=3)
        self.height_label = tk.Label(self.root, text="图片长度:")
        self.height_label.grid(row=1, column=3)
        self.author_label = tk.Label(self.root, text="by William")
        self.author_label.place(relx=1.0, rely=1.0, anchor='se')
        # 文本
        self.list_text = tk.Text(self.root, width=50, height=1)  # 截图名单录入框
        self.list_text.grid(row=0, column=1, rowspan=1, columnspan=1)
        self.folder_text = tk.Text(self.root, width=50, height=1)  # 输出目录
        self.folder_text.grid(row=1, column=1, rowspan=1, columnspan=1)
        self.info_text = tk.Text(self.root, width=50, height=20)  # 运行信息
        self.info_text.grid(row=10, column=1, rowspan=10, columnspan=1)
        self.width_text = tk.Text(self.root, width=5, height=1)  # 宽度设置
        self.width_text.grid(row=0, column=4, rowspan=1, columnspan=1)
        self.height_text = tk.Text(self.root, width=5, height=1)  # 高度设置
        self.height_text.grid(row=1, column=4, rowspan=1, columnspan=1)

        # 按钮
        self.list_button = tk.Button(self.root,
                                     text="选择名单",
                                     bg="white",
                                     width=10,
                                     height=1,
                                     command=self.get_file_path)  # 调用内部方法,加()为直接调用
        self.list_button.grid(row=0, column=2)
        self.folder_button = tk.Button(self.root,
                                       text="指定目录",
                                       bg="white",
                                       width=10,
                                       height=1,
                                       command=self.get_folder_path)  # 调用内部方法,加()为直接调用
        self.folder_button.grid(row=1, column=2)
        self.folder_button = tk.Button(self.root,
                                       text="文件检查",
                                       bg="white",
                                       width=10,
                                       height=1,
                                       command=self.get_info)  # 调用内部方法,加()为直接调用
        self.folder_button.grid(row=10, column=2)
        self.exe_button = tk.Button(self.root,
                                    text="开始截图",
                                    bg="white",
                                    width=10,
                                    height=1,
                                    command=self.exe)  # 调用内部方法,加()为直接调用
        self.exe_button.grid(row=11, column=2)
        self.default_button = tk.Button(self.root,
                                        text="设为默认",
                                        bg="white",
                                        width=10,
                                        height=1,
                                        command=self.set_default)  # 调用内部方法,加()为直接调用
        self.default_button.grid(row=10, column=3, columnspan=2)

        self.set_default()

    def set_default(self):
        self.width_text.delete(1.0, 'end')
        self.height_text.delete(1.0, 'end')
        self.info_text.delete(1.0, 'end')
        self.width_text.insert(1.0, 900)
        self.height_text.insert(1.0, 2250)

        photo = tk.PhotoImage(file='format.gif')
        photo_label = tk.Label(image=photo)
        photo_label.image = photo

        self.info_text.insert(1.0, '欢迎使用\n\n请将文件按照以下格式编辑：')
        self.info_text.image_create(10.0, image=photo)

    def get_file_path(self):
        self.list_text.delete(1.0, 'end')
        self.filepath = filedialog.askopenfilename()  # 获得选择好的文件
        self.list_text.insert(1.0, self.filepath)

    def get_folder_path(self):
        self.folder_text.delete(1.0, 'end')
        self.folderpath = filedialog.askdirectory()  # 获得选择好的文件夹
        self.folder_text.insert(1.0, self.folderpath)

    def get_info(self):
        self.filepath = self.list_text.get("1.0", "end").strip()
        self.folderpath = self.folder_text.get("1.0", "end").strip()

        try:
            self.name_list = read_cols(self.filepath, "A")[1:]
            self.url_list = read_cols(self.filepath, "B")[1:]
            self.check = True
            if not os.path.exists(self.folderpath):
                tk.messagebox.showerror('错误', '请提供正确的目标文件夹')
                self.check = False

        except Exception:
            self.check = False
            tk.messagebox.showerror('错误', '请提供正确的文件路径')

        self.info_text.delete(1.0, 'end')

        if os.path.exists(self.folderpath):
            self.info_text.insert(1.0, '输出路径: \n' + self.folderpath + '\n\n')
        else:
            self.info_text.insert(1.0, '错误的输出文件夹\n')

        if os.path.isfile(self.filepath):
            self.info_text.insert(1.0, '文件路径: \n' + self.filepath + '\n\n')
        else:
            self.info_text.insert(1.0, '错误的文件路径\n')

        try:
            self.info_text.insert(1.0, '获得名单: \n' +
                                  self.name_list[0] + '\n' +
                                  self.name_list[1] + '\n' +
                                  self.name_list[2] + '\n' +
                                  '等，共计' + str(len(self.name_list)) + '个\n\n')

            self.info_text.insert(1.0, '获得地址: \n' +
                                  self.url_list[0] + '\n' +
                                  self.url_list[1] + '\n' +
                                  self.url_list[2] + '\n' +
                                  '等，共计' + str(len(self.url_list)) + '个\n\n')
        except Exception:
            self.info_text.insert(1.0, '获得名单: \n' +
                                  str(self.name_list) +  '\n' +
                                  '等，共计' + str(len(self.name_list)) + '个\n\n')
            self.info_text.insert(1.0, '获得地址: \n' +
                                  str(self.url_list) +  '\n' +
                                  '等，共计' + str(len(self.url_list)) + '个\n\n')

        self.number_check()
        if self.check:
            tk.messagebox.showinfo('就绪', '一切准备就绪\n可以开始截图')

    def exe(self):
        if self.check:
            self.screenshot()
        else:
            tk.messagebox.showerror("错误", '请先检查文件')

    def screenshot(self):
        names = []
        urls = []
        for i in self.name_list:
            names.append(i)
        for i in self.url_list:
            urls.append(i)
        number = 1

        width = self.width_text.get("1.0", "end").strip()
        height = self.height_text.get("1.0", "end").strip()

        while urls:

            url = urls.pop(0)
            name = str(names.pop(0))
            self.info_text.insert(1.0, '正在对' + name + '的笔记进行截图，这是第' + str(
                number) + '个\n\n')
            self.root.update()

            try:
                self.auto_shots(url, name, width, height)
                self.info_text.insert(1.0,
                                      "✅ %s：截图完毕," % name + '第' + str(number) + '个已完成\n\n')
                self.success_count += 1
            except Exception:
                self.info_text.insert(1.0,
                                      "❎ %s：截图失败," % name + '第' + str(number) + '个未完成\n❎ 警告：网址不可用\n\n')
                self.fail_count += 1
                self.fail_list.append(name)


            self.root.update()
            number = number + 1

        self.info_text.insert(1.0, '全部截图已完成，共计' + str(len(self.name_list)) + '个\n\n')
        self.output_txt()
        tk.messagebox.showinfo('已完成', '所有笔记都截好啦！！')

    def auto_shots(self, url, name, width, height):
        chromedriver = "chromedriver.exe"
        os.environ["webdriver.chrome.driver"] = chromedriver
        chrome_options = Options()
        chrome_options.add_argument('headless')
        driver = webdriver.Chrome(chromedriver, options=chrome_options)

        picture_url = url
        driver.get(picture_url)
        time.sleep(2)

        driver.set_window_size(width, height)

        file_name = self.folderpath + "\\" + name + ".png"
        driver.save_screenshot(file_name)
        driver.close()

    def number_check(self):
        status = False
        if self.width_text.get("1.0", "end").strip() == '':
            status = True
        if int(self.width_text.get("1.0", "end").strip()) < 0:
            status = True
        if int(self.width_text.get("1.0", "end").strip()) > 3000:
            status = True
        if self.height_text.get("1.0", "end").strip() == '':
            status = True
        if int(self.height_text.get("1.0", "end").strip()) < 0:
            status = True
        if int(self.height_text.get("1.0", "end").strip()) > 3000:
            status = True
        if status:
            self.set_default()

    def output_txt(self):
        localtime = time.localtime(time.time())
        tttt = time.strftime('%Y:%m:%d %H:%M:%S', localtime).replace(':', '-', 2).replace(':', '_')
        txt_name = self.folderpath + "\\截图日志" + tttt + '.txt'
        txt_in_progress = open(txt_name, mode='w')

        total = len(self.name_list)
        lines_to_write = []
        lines_to_write.append('截图总数：' + str(total) + '\n')
        lines_to_write.append('其中成功' + str(self.success_count) +
                              '个，失败' + str(self.fail_count) + '个。\n\n')
        if self.fail_list:
            lines_to_write.append('以下为失败名单：\n')
            for i in self.fail_list:
                lines_to_write.append(i + '\n')

        txt_in_progress.writelines(lines_to_write)
        txt_in_progress.close()


def gui_start():
    root = tk.Tk()  # 实例化出一个父窗口

    portal = GUI(root)
    portal.set()

    root.mainloop()  # 父窗口进入事件循环，可以理解为保持窗口运行，否则界面不展示


def read_cols(file_name, col_name):

    table_1 = load_workbook(file_name)
    sheets = []
    for i in table_1.sheetnames:
        sheets.append(i)
    sheet_1 = table_1[sheets[0]]
    result = []
    for cell in sheet_1[col_name]:
        result.append(cell.value)
    while result[-1] is None:
        result.pop()

    return result


if __name__ == "__main__":
    gui_start()
