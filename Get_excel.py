from openpyxl import load_workbook

from datetime import date, datetime


def read_cols(file_name, col_name):

    table_1 = load_workbook(file_name)
    sheets = []
    for i in table_1.sheetnames:
        sheets.append(i)

    sheet_1 = table_1[sheets[0]]

    result = []

    for cell in sheet_1[col_name]:
        result.append(cell.value)

    while result[-1] is None:
        result.pop()

    return result


if __name__ == "__main__":

    xlsx_path = r"test2.xlsx"
    name_list = read_cols(xlsx_path, "A")
    url_list = read_cols(xlsx_path, "B")

    print(name_list)
    print(url_list)
    print(len(url_list) == len(name_list))
    print(len(url_list))
